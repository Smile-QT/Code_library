
# 批量爬取抖音视频的文案

from selenium import webdriver
import xlrd
import openpyxl as op

# #  #
# //*[@id="root"]/div/div[2]/div/div/div[1]/div[3]/div/div[1]/div/h1/span[2]/span[1]/a/span
# //*[@id="root"]/div/div[2]/div/div/div[1]/div[3]/div/div[1]/div/h1/span[2]/span[2]/a/span
#
# # 文案
# //*[@id="root"]/div/div[2]/div/div/div[1]/div[3]/div/div[1]/div/h1/span[2]/span[4]/span/span/span
# //*[@id="root"]/div/div[2]/div/div/div[1]/div[3]/div/div[1]/div/h1/span[2]/span[1]/span/span/span
# //*[@id="root"]/div/div[2]/div/div/div[1]/div[3]/div/div[1]/div/h1/span[2]/span[4]/span/span/span


#########################################################################################################################################
# 读取抖音视频的网址

filePath = r"C:/YXL/lijiawen/reptile/wenan/Data/xinnongrenjihua02.xlsx"
data = xlrd.open_workbook(filePath)

sheet1 = data.sheet_by_name("Sheet1")   # 按名获取该文件中的表格

zero_col_values = sheet1.col_values(0)  # 获取第1列的数据：抖音视频的网址的序号，序号是从1开始的。
# 将字符串数字，转化成int型的数字，方便之后使用
for i in range(1, len(zero_col_values)):
    zero_col_values[i] = int(zero_col_values[i])

first_col_values = sheet1.col_values(1)  # 获取第2列的数据：抖音视频的网址


for j in range(1, len(first_col_values)):
    driver = webdriver.Chrome('C:/Users/admin/anaconda3/envs/pytorchYXL/chromedriver_win32_v1/chromedriver.exe')  # chrome驱动程序存放的位置

    douying_vedio_path = first_col_values[j]
    douyingshipingxuhao_pinglunsuoxiehangshu = zero_col_values[j]

    driver.get(douying_vedio_path)  #网页版抖音的网址

    for cunt in range(1, 10):

        # 提取#
        jinhao = driver.find_element_by_xpath('//*[@id="root"]/div/div[2]/div/div/div[1]/div[3]/div/div[1]/div/h1/span[2]/span[' + str(cunt) + ']/a/span')  # 需要爬的#

        # 当有#数据时才写进文件
        if len(str(jinhao.text)) > 0:
            # 这是把爬下来的评论存放在文件中
            tableAll = op.load_workbook(filePath)
            table1 = tableAll['Sheet1']
            table1.cell(douyingshipingxuhao_pinglunsuoxiehangshu + 1, 3, str(jinhao.text))
            tableAll.save(filePath)
            print(1)


        # 提取文案
        wenan = driver.find_element_by_xpath('//*[@id="root"]/div/div[2]/div/div/div[1]/div[3]/div/div[1]/div/h1/span[2]/span['+str(cunt) +']/span/span/span')  # 需要爬的文案


        # 当有文案数据时才写进文件
        if len(str(wenan.text)) > 0:
            # 这是把爬下来的评论存放在文件中
            tableAll = op.load_workbook(filePath)
            table1 = tableAll['Sheet1']
            table1.cell(douyingshipingxuhao_pinglunsuoxiehangshu + 1, 4, str(wenan.text))
            tableAll.save(filePath)





